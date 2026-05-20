from openpyxl import load_workbook
import sys,io
sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
for path in [r'c:\Projects\CLIENTS\CHEEL4REEL\інше\Loyalty_Program_Adjustment_C4R.xlsx', r'c:\Projects\CLIENTS\CHEEL4REEL\інше\FS - Games.xlsx']:
    print('='*80); print('FILE:', path); print('='*80)
    wb = load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'\n--- SHEET: {sn} (rows={ws.max_row} cols={ws.max_column}) ---')
        for row in ws.iter_rows(values_only=True):
            print(' | '.join('' if v is None else str(v) for v in row))
