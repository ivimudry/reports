"""Fix tiered emails in HeyHo Emails Summary.xlsx"""
import openpyxl

XLSX = r"C:\Projects\REPORTS\тексти\хейхо\HeyHo Emails Summary.xlsx"
wb = openpyxl.load_workbook(XLSX)

ws = wb['Nutrition #1 SLOTS']

# Find rows by email name (column A, data starts row 3)
fixes = {
    'Email 2.1': {
        'bonus': '30/50/70 FS (tiered)',
        'promo': 'ASKFORMORE1 / ASK4MOREV7X / ASKFORMZKQ2',
        'game': 'Reactoonz 2',
    },
    'Email 2.2': {
        'bonus': '50/70/90 FS (tiered)',
        'promo': 'ASK4MOREV7X / ASKFORMZKQ2 / LEGA90',
        'game': 'Reactoonz 2',
    },
    'Email 2.3': {
        'bonus': '70/90/120 FS (tiered)',
        'promo': 'ASKFORMZKQ2 / LEGA90 / CROWN120',
        'game': 'Reactoonz 2',
    },
    'Email 4.1': {
        'bonus': '30/50/70 FS (tiered)',
        'promo': 'ASKFORMORE1 / ASK4MOREV7X / ASKFORMZKQ2',
        'game': '27 Dice',
    },
    'Email 4.2': {
        'bonus': '50/70/90 FS (tiered)',
        'promo': 'ASK4MOREV7X / ASKFORMZKQ2 / LEGA90',
        'game': '27 Dice',
    },
    'Email 4.3': {
        'bonus': '70/90/120 FS (tiered)',
        'promo': 'ASKFORMZKQ2 / LEGA90 / CROWN120',
        'game': '27 Dice',
    },
}

for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if name in fixes:
        f = fixes[name]
        ws.cell(r, 2, f['bonus'])
        ws.cell(r, 3, f['promo'])
        ws.cell(r, 4, f['game'])
        print(f"Fixed: {name} -> {f['bonus']} | {f['promo']} | {f['game']}")

wb.save(XLSX)
print("\nSaved!")
