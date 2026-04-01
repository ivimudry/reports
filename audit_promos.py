"""Audit: compare promo codes from HeyHo Emails Summary vs Promo codes master table."""

import openpyxl
import re
import sys

PROMO_FILE = r"C:\Projects\REPORTS\тексти\хейхо\Promo codes - HeyHo Casino.xlsx"
EMAILS_FILE = r"C:\Projects\REPORTS\тексти\хейхо\HeyHo Emails Summary.xlsx"

# Mapping: email summary sheet -> promo code master sheet(s)
# HH_Live is one sheet for all LIVE campaigns
SHEET_MAP = {
    'Welcome Flow': 'Welcome Flow',
    'DEP Retention': 'DEP Retention',
    'SU Retention': 'SU Retention',
    'Nutrition #1 SLOTS': 'Nutrition #1',
    'Nutrition #1 LIVE': 'HH_Live',
    'Nutrition #2 SLOTS': 'Nutrition #2',
    'Nutrition #2 LIVE': 'HH_Live',
    'Nutrition #3 SLOTS': 'Nutrition #3',
    'Nutrition #3 LIVE': 'HH_Live',
    'Nutrition #4 SLOTS': 'Nutrition #4',
    'Nutrition #4 LIVE': 'HH_Live',
    'Nutrition #5 SLOTS': 'Nutrition #5',
    'Nutrition #5 LIVE': 'HH_Live',
    'Nutrition #6 SLOTS': 'Nutrition #6',
    'Nutrition #6 LIVE': 'HH_Live',
    'Unsuccessful Deposit': 'Nutrition #2',  # GIVEMEMORELZ8 is from Nutrition #2
}


def load_promo_master(filepath):
    """Load all promo codes from the master table. Returns dict: sheet -> {code: {offer, game}}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        codes = {}
        
        if sn == 'HH_Live':
            # Columns: #, Promo Code, Deposit#, Bonus Desired, Bonus Actual, T&C, Wager
            for r in range(2, ws.max_row + 1):
                code = str(ws.cell(r, 2).value or '').strip()
                deposit_num = ws.cell(r, 3).value
                bonus_desired = ws.cell(r, 4).value
                if code and code != 'Промо код':
                    offer_str = ''
                    if bonus_desired is not None:
                        try:
                            val = float(bonus_desired)
                            offer_str = f"{int(val*100)}%"
                        except (ValueError, TypeError):
                            offer_str = str(bonus_desired)
                    codes[code] = {'offer': offer_str, 'game': '', 'deposit': deposit_num}
        else:
            # Standard columns: #, Promo Code, Offer, Game, ...
            code_col = 2
            offer_col = 3
            game_col = 4
            
            # For Aff_bonuses, columns are shifted
            if sn == 'Aff_bonuses':
                code_col = 3
                offer_col = 4
                game_col = 5
            
            for r in range(2, ws.max_row + 1):
                code = str(ws.cell(r, code_col).value or '').strip()
                offer = str(ws.cell(r, offer_col).value or '').strip()
                game = str(ws.cell(r, game_col).value or '').strip()
                if code and code not in ('Promo Code', 'Промо код', 'Bonus name', 'None', ''):
                    codes[code] = {'offer': offer, 'game': game}
        
        result[sn] = codes
    
    wb.close()
    return result


def load_email_summary(filepath):
    """Load all emails from summary. Returns dict: sheet -> [{name, bonus, promo, game}]"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    
    for sn in wb.sheetnames:
        if sn == 'Summary':
            continue
        ws = wb[sn]
        emails = []
        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            emails.append({
                'name': str(name),
                'bonus': str(ws.cell(r, 2).value or ''),
                'promo': str(ws.cell(r, 3).value or '').strip(),
                'game': str(ws.cell(r, 4).value or '').strip(),
            })
        result[sn] = emails
    
    wb.close()
    return result


def normalize_offer(offer_str):
    """Normalize offer string for comparison."""
    s = offer_str.strip()
    # Convert decimal like "1.0" to "100%", "1.5" to "150%", etc.
    try:
        val = float(s)
        if 0 < val <= 10:
            return f"{int(val*100)}%"
    except ValueError:
        pass
    
    s = s.replace('free spins', 'FS').replace('Free Spins', 'FS')
    s = s.replace(' (quick bonus)', '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def compare():
    print("Loading promo master table...")
    master = load_promo_master(PROMO_FILE)
    
    # Build global code->sheet mapping to detect cross-campaign usage
    code_to_campaigns = {}
    for sn, codes in master.items():
        for code in codes:
            code_to_campaigns.setdefault(code, []).append(sn)
    
    print("Loading email summary...")
    emails = load_email_summary(EMAILS_FILE)
    
    issues = []
    ok_count = 0
    
    for sheet_name, email_list in emails.items():
        master_sheet = SHEET_MAP.get(sheet_name)
        if not master_sheet:
            issues.append(f"[?] No master sheet mapping for: {sheet_name}")
            continue
        
        master_codes = master.get(master_sheet, {})
        
        # Also build a "all codes for this campaign context" lookup
        # For HH_Live, all live codes are in one sheet
        
        print(f"\n{'='*60}")
        print(f"  {sheet_name} -> master: {master_sheet} ({len(master_codes)} codes)")
        print(f"{'='*60}")
        
        for em in email_list:
            promo = em['promo']
            if not promo or promo == 'None':
                continue
            
            # Check 1: Is promo code in the correct master sheet?
            if promo in master_codes:
                master_info = master_codes[promo]
                m_offer = normalize_offer(master_info['offer'])
                e_bonus = em['bonus']
                
                # Check bonus match
                bonus_match = True
                
                # For HH_Live, the offer is a percentage
                if master_sheet == 'HH_Live':
                    # Extract percentage from email bonus
                    pct_match = re.search(r'(\d+)%', e_bonus)
                    if pct_match and m_offer:
                        e_pct = pct_match.group(1) + '%'
                        if m_offer == 'Cashback':
                            pass  # Cashback codes are in DEP Retention, not HH_Live
                        elif e_pct != m_offer:
                            bonus_match = False
                            issues.append(f"[BONUS MISMATCH] {sheet_name} | {em['name']} | code={promo} | email says: {e_bonus} | master says: {m_offer}")
                else:
                    # For slots: compare FS and % numbers
                    m_fs = re.search(r'(\d+)\s*FS', m_offer)
                    e_fs = re.search(r'(\d+)\s*FS', e_bonus)
                    m_pct = re.search(r'(\d+)%', m_offer)
                    e_pct = re.search(r'(\d+)%', e_bonus)
                    m_cb = 'CB' in m_offer.upper() or 'CASHBACK' in m_offer.upper()
                    e_cb = 'cashback' in e_bonus.lower() or 'cash back' in e_bonus.lower()
                    
                    if m_fs and e_fs:
                        if m_fs.group(1) != e_fs.group(1):
                            bonus_match = False
                            issues.append(f"[FS MISMATCH] {sheet_name} | {em['name']} | code={promo} | email: {e_bonus} | master: {m_offer}")
                    if m_pct and e_pct and not m_cb and not e_cb:
                        if m_pct.group(1) != e_pct.group(1):
                            bonus_match = False
                            issues.append(f"[% MISMATCH] {sheet_name} | {em['name']} | code={promo} | email: {e_bonus} | master: {m_offer}")
                
                # Check game match
                m_game = master_info.get('game', '').strip()
                e_game = em['game'].strip()
                if m_game and e_game and m_game.lower() != e_game.lower():
                    issues.append(f"[GAME MISMATCH] {sheet_name} | {em['name']} | code={promo} | email game: {e_game} | master game: {m_game}")
                
                if bonus_match:
                    ok_count += 1
                    status = 'OK'
                else:
                    status = 'MISMATCH'
                
                print(f"  {status:10s} {em['name']:15s} | {promo:15s} | email: {e_bonus:30s} | master: {m_offer}")
            
            else:
                # Code NOT in the expected master sheet - check where it belongs
                found_in = code_to_campaigns.get(promo, [])
                if found_in:
                    issues.append(f"[WRONG CAMPAIGN] {sheet_name} | {em['name']} | code={promo} | NOT in {master_sheet}, found in: {', '.join(found_in)}")
                    print(f"  {'WRONG_CAMP':10s} {em['name']:15s} | {promo:15s} | NOT in {master_sheet}, found in: {', '.join(found_in)}")
                else:
                    issues.append(f"[MISSING CODE] {sheet_name} | {em['name']} | code={promo} | NOT found in ANY master sheet")
                    print(f"  {'MISSING':10s} {em['name']:15s} | {promo:15s} | NOT found in any master sheet")
    
    # Summary
    print(f"\n{'='*60}")
    print(f"AUDIT SUMMARY")
    print(f"{'='*60}")
    print(f"OK: {ok_count}")
    print(f"Issues: {len(issues)}")
    
    if issues:
        print(f"\nALL ISSUES:")
        for i, iss in enumerate(issues, 1):
            print(f"  {i}. {iss}")


if __name__ == '__main__':
    compare()
