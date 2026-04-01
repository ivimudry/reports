"""
1. LIVE audit with CORRECT master bonuses (from screenshot)
2. Check FS-without-game emails for game names in source .txt files
"""

import openpyxl, re, os

SUMMARY = os.path.join("тексти", "хейхо", "HeyHo Emails Summary.xlsx")
TXT_DIR = os.path.join("тексти", "хейхо")

# ── Correct HH_Live master from screenshot ──────────────────────────
LIVE_MASTER = {
    "HOOK150":  "130%",
    "RAID145":  "125%",
    "MAP140":   "120%",
    "SABR135":  "115%",
    "KRKN130":  "110%",
    "RUM125":   "105%",
    "CREW120":  "100%",
    "FLAG115":  "95%",
    "KEEL110":  "90%",
    "BUCC105":  "85%",
    "SKLL100":  "80%",
    "REEF90":   "75%",
    "PLANK80":  "70%",
    "CHEST75":  "65%",
    "CUTLS70":  "60%",
    "CB4RUM":   "4% CB",
    "CB6MAP":   "6% CB",
    "CB9REEF":  "9% CB",
    "CB12SEA":  "12% CB",
    "CB15BAY":  "15% CB",
}

# Also note deposit groups from screenshot:
# Dep 2: HOOK150..CREW120 (rows 1-7)
# Dep 3,4: FLAG115..SKLL100 (rows 8-11)
# Dep 5,6: REEF90..CUTLS70 (rows 12-15)
# CB: CB4RUM..CB15BAY (rows 16-20)

# ═══════════════════════════════════════════════════════════════════
# PART 1: LIVE audit
# ═══════════════════════════════════════════════════════════════════
print("=" * 70)
print("PART 1: LIVE campaigns vs HH_Live master (CORRECT bonuses)")
print("=" * 70)

wb = openpyxl.load_workbook(SUMMARY, data_only=True)
live_sheets = [s for s in wb.sheetnames if "LIVE" in s.upper()]

live_ok = 0
live_issues = []

for sname in live_sheets:
    ws = wb[sname]
    print(f"\n--- {sname} ---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        email = str(row[0]).strip()
        bonus = str(row[1]).strip() if row[1] else ""
        code  = str(row[2]).strip() if row[2] else ""
        game  = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if email == "Email" or code == "Promo Code":
            continue

        code_upper = code.upper()
        if not code_upper or code_upper not in LIVE_MASTER:
            if not code_upper:
                # Info emails (no promo code) - just note them
                print(f"  {email:25s} | {bonus[:40]:40s} | (no code - info email)")
            else:
                live_issues.append((sname, email, f"Code '{code}' NOT in master"))
                print(f"  {email:25s} | code={code} NOT FOUND")
            continue

        master_bonus = LIVE_MASTER[code_upper]

        # Normalize
        email_norm = bonus.lower().replace(" ", "").replace("cashback", "cb")
        master_norm = master_bonus.lower().replace(" ", "")

        # Extract % for comparison
        email_pct = re.search(r'(\d+)%', email_norm)
        master_pct = re.search(r'(\d+)%', master_norm)

        match = False
        if email_pct and master_pct:
            match = email_pct.group(1) == master_pct.group(1)

        if match:
            live_ok += 1
            print(f"  {email:25s} | {bonus:20s} | {code:12s} | OK")
        else:
            live_issues.append((sname, email, f"email='{bonus}' vs master='{master_bonus}' (code={code})"))
            print(f"  {email:25s} | {bonus:20s} | {code:12s} | *** MISMATCH *** master={master_bonus}")

print(f"\n{'='*50}")
print(f"LIVE Summary: OK={live_ok}, Issues={len(live_issues)}")
if live_issues:
    for s, e, msg in live_issues:
        print(f"  [{s}] {e}: {msg}")
else:
    print("  All LIVE bonuses match master!")

# ═══════════════════════════════════════════════════════════════════
# PART 2: Check FS-without-game emails in source .txt files
# ═══════════════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("PART 2: Checking FS-without-game emails for game names in source text")
print("=" * 70)

# Emails to check:
# N1 SLOTS: Email 6.3, Email 6.8
# N5 SLOTS: Emails 4.1-4.8
# Welcome Flow: Emails 1, 2, 4

KNOWN_GAMES = [
    "Sugar Rush", "Sweet Bonanza", "Gates of Olympus", "Big Bass",
    "Starlight Princess", "Dog House", "Book of Dead", "Reactoonz",
    "Starburst", "Gonzos Quest", "Gonzo's Quest", "Fire Joker",
    "Wolf Gold", "Mustang Gold", "John Hunter", "Release the Kraken",
    "Madame Destiny", "Great Rhino", "Buffalo King", "Fruit Party",
    "Wanted Dead", "Hand of Anubis", "Floating Dragon", "5 Lions",
    "Cleocatra", "Plinko", "Aviator", "Space XY", "JetX", "Jet X",
    "Mines", "Tome of Madness", "Rise of Merlin", "Legacy of Dead",
    "Book of Fallen", "Joker King", "Extra Chilli", "Bonanza",
    "Razor Shark", "Wild West Gold", "Mental", "Cash Bonanza",
    "Money Train", "Might of Ra", "Zeus vs Hades", "Wisdom of Athena",
    "27 Dice", "Hot Fruits", "Cash Elevator", "Fire Strike",
]

checks = [
    ("Nutrition #1 SLOTS - Table data.txt", ["Email 6.3", "Email 6.8"]),
    ("Nutrition #5 SLOTS - Table data.txt", [f"Email 4.{i}" for i in range(1, 9)]),
    ("Welcome Flow - Table data.txt", ["Email 1", "Email 2", "Email 4"]),
]

game_fixes = []  # (sheet_name, email_name, game_found)

for filename, email_names in checks:
    fpath = os.path.join(TXT_DIR, filename)
    if not os.path.exists(fpath):
        print(f"\n  FILE NOT FOUND: {filename}")
        continue

    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into blocks
    blocks = re.split(r'={5,}', content)

    for target_email in email_names:
        print(f"\n--- {filename} / {target_email} ---")
        found_block = None

        for block in blocks:
            block_stripped = block.strip()
            if not block_stripped:
                continue
            # Check if this block contains the target email name
            # Look for "Email name: XXX" or similar
            name_match = re.search(
                r'(?:Email\s*name|name)\s*:\s*(.+)',
                block_stripped, re.IGNORECASE
            )
            if name_match:
                bname = name_match.group(1).strip()
                # For Welcome Flow emails, names are like "HH_WF_Email 1 [Default]"
                # For SLOTS, names are like "HH_N1S_Email 6.3 [Default]"
                # We need to match flexibly
                if target_email.replace(" ", "").lower() in bname.replace(" ", "").lower():
                    # Only take Default locale block
                    if "Default" in block_stripped or "default" in block_stripped:
                        found_block = block_stripped
                        break

        if not found_block:
            # Try looser match - just find block with the email number
            for block in blocks:
                block_stripped = block.strip()
                if not block_stripped:
                    continue
                # Check for email name pattern
                email_num = target_email.replace("Email ", "")
                pattern = re.compile(
                    rf'(?:Email\s*name|name)\s*:\s*.*?{re.escape(email_num)}\b.*?\[Default\]',
                    re.IGNORECASE
                )
                if pattern.search(block_stripped):
                    found_block = block_stripped
                    break

        if not found_block:
            # Even looser: find any block mentioning the email
            for block in blocks:
                block_stripped = block.strip()
                email_num = target_email.replace("Email ", "")
                if re.search(rf'Email\s*{re.escape(email_num)}\b', block_stripped) and '[Default]' in block_stripped:
                    found_block = block_stripped
                    break

        if found_block:
            # Print full block for review
            print(f"  [FOUND - showing full text fields]")

            # Extract all text fields
            text_fields = []
            for line in found_block.split('\n'):
                line_s = line.strip()
                if ':' in line_s:
                    key = line_s.split(':')[0].strip().lower()
                    val = line_s.split(':', 1)[1].strip()
                    if re.match(r'text_\d+|subject|preheader', key):
                        text_fields.append((key, val))

            for key, val in text_fields:
                print(f"    {key}: {val}")

            # Search for game names in the block
            found_games = []
            block_lower = found_block.lower()
            for game in KNOWN_GAMES:
                if game.lower() in block_lower:
                    found_games.append(game)

            # Also search for common game patterns
            # e.g. "on [Game Name]" or "in [Game Name]" or "play [Game Name]"
            game_pattern_matches = re.findall(
                r'(?:play|on|in|at|slot|game|spin|spins on|Free Spins on|FS on)\s+([A-Z][A-Za-z0-9\s\':]+?)(?:\s*[!.,;:\-\n]|$)',
                found_block
            )
            if game_pattern_matches:
                print(f"  Potential game references: {game_pattern_matches}")

            if found_games:
                print(f"  >>> GAMES FOUND: {found_games}")
                # Map filename to sheet name
                sheet_map = {
                    "Nutrition #1 SLOTS - Table data.txt": "Nutrition #1 SLOTS",
                    "Nutrition #5 SLOTS - Table data.txt": "Nutrition #5 SLOTS",
                    "Welcome Flow - Table data.txt": "Welcome Flow",
                }
                sheet_name = sheet_map.get(filename, filename.replace(" - Table data.txt", ""))
                game_fixes.append((sheet_name, target_email, found_games[0]))
            else:
                print(f"  >>> No game names found in text")
        else:
            print(f"  [BLOCK NOT FOUND]")

# Summary
print(f"\n{'='*50}")
print("Game fixes to apply:")
if game_fixes:
    for sheet, email, game in game_fixes:
        print(f"  [{sheet}] {email} -> game: {game}")
else:
    print("  No game names found in any of the checked emails.")

wb.close()
