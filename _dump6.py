from openpyxl import load_workbook
path = r"c:\Projects\CLIENTS\CHEEL4REEL\інше\Chill4Reel_Welcome_Pack&Weekly_Bonuses_Consulting_UPD_APR2026.xlsx"
wb = load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    maxcol = 0
    for row in ws.iter_rows(values_only=True):
        for j,v in enumerate(row):
            if v is not None and str(v).strip(): maxcol = max(maxcol, j+1)
    print(f"\n========== SHEET: {sn} (maxcol={maxcol}, maxrow={ws.max_row}) ==========")
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        trimmed = row[:maxcol]
        if not any(v is not None and str(v).strip() for v in trimmed): continue
        s = " | ".join("" if v is None else str(v).replace("\n"," / ").strip() for v in trimmed)
        print(f"R{i}: {s}")
