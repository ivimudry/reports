from openpyxl import load_workbook
import sys
sys.stdout.reconfigure(encoding='utf-8')

print("="*80); print("FS - Games.xlsx"); print("="*80)
wb = load_workbook(r"c:\Projects\CLIENTS\CHEEL4REEL\інше\FS - Games.xlsx", data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"--- {sn} ---")
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        print(f"R{i}: " + " | ".join("" if v is None else str(v) for v in row))

print("\n"+"="*80); print("Loyalty_Program_Adjustment_C4R.xlsx"); print("="*80)
wb = load_workbook(r"c:\Projects\CLIENTS\CHEEL4REEL\інше\Loyalty_Program_Adjustment_C4R.xlsx", data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"--- {sn} ---")
    maxcol = 0
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for j,v in enumerate(row):
            if v is not None and str(v).strip(): maxcol = max(maxcol, j+1)
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        trimmed = row[:maxcol] if maxcol else row
        if not any(v is not None and str(v).strip() for v in trimmed): continue
        print(f"R{i}: " + " | ".join("" if v is None else str(v) for v in trimmed))
