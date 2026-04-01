"""Re-audit LIVE campaigns vs updated HH_Live master table,
   and scan ALL campaigns for multi-bonus emails."""

import openpyxl, re, os

SUMMARY = os.path.join("тексти", "хейхо", "HeyHo Emails Summary.xlsx")
MASTER  = os.path.join("тексти", "хейхо", "Promo codes - HeyHo Casino.xlsx")

# ── 1. Read updated HH_Live master table ────────────────────────────
wb_m = openpyxl.load_workbook(MASTER, data_only=True)
ws_live = wb_m["HH_Live"]

print("=== HH_Live master table (updated) ===")
master = {}  # promo_code -> {bonus, deposit, ...}
for row in ws_live.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    num, code, deposit, bonus_desired, bonus_actual, tc, wager = (
        list(row) + [None]*7)[:7]
    code_str = str(code).strip() if code else ""
    if not code_str:
        continue
    master[code_str.upper()] = {
        "code": code_str,
        "deposit": str(deposit).strip() if deposit else "",
        "bonus_desired": str(bonus_desired).strip() if bonus_desired else "",
        "bonus_actual": str(bonus_actual).strip() if bonus_actual else "",
    }
    print(f"  {code_str:20s} | dep={deposit} | desired={bonus_desired} | actual={bonus_actual}")

print(f"\nTotal master codes: {len(master)}\n")

# ── 2. Read LIVE sheets from summary ────────────────────────────────
wb_s = openpyxl.load_workbook(SUMMARY, data_only=True)

live_sheets = [s for s in wb_s.sheetnames if "LIVE" in s.upper()]
print(f"LIVE sheets found: {live_sheets}\n")

issues = []
ok_count = 0

for sname in live_sheets:
    ws = wb_s[sname]
    print(f"--- {sname} ---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        email_name = str(row[0]).strip()
        bonus = str(row[1]).strip() if row[1] else ""
        promo = str(row[2]).strip() if row[2] else ""
        game  = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        promo_upper = promo.upper()
        if promo_upper in master:
            m = master[promo_upper]
            # Compare bonus
            # Try to find the bonus % in the email bonus string
            print(f"  {email_name:30s} | bonus={bonus:25s} | code={promo:20s} | master_desired={m['bonus_desired']:15s} | master_actual={m['bonus_actual']}")

            # Simple check: does the master bonus appear somewhere in the email bonus?
            desired = m["bonus_desired"]
            actual = m["bonus_actual"]

            # Extract numbers from bonus strings for comparison
            email_nums = set(re.findall(r'\d+', bonus))
            desired_nums = set(re.findall(r'\d+', desired))
            actual_nums = set(re.findall(r'\d+', actual))

            # Check if email bonus matches either desired or actual
            match = False
            if desired_nums and desired_nums.issubset(email_nums):
                match = True
            if actual_nums and actual_nums.issubset(email_nums):
                match = True
            # Also check direct string containment
            if desired and desired in bonus:
                match = True
            if actual and actual in bonus:
                match = True

            if match:
                ok_count += 1
            else:
                issues.append((sname, email_name, bonus, promo, m["bonus_desired"], m["bonus_actual"]))
                print(f"    *** MISMATCH ***")
        else:
            issues.append((sname, email_name, bonus, promo, "NOT FOUND", ""))
            print(f"  {email_name:30s} | code={promo} NOT FOUND in master")

print(f"\n=== LIVE Audit Summary ===")
print(f"OK: {ok_count}")
print(f"Issues: {len(issues)}")
for sname, email, bonus, code, desired, actual in issues:
    print(f"  [{sname}] {email} | email_bonus={bonus} | code={code} | master_desired={desired} | master_actual={actual}")

# ── 3. Scan ALL campaign .txt files for multi-bonus emails ──────────
print("\n\n=== Scanning for multi-bonus emails across ALL campaigns ===\n")

txt_dir = os.path.join("тексти", "хейхо")
multi_bonus = []

for fname in sorted(os.listdir(txt_dir)):
    if not fname.endswith(".txt"):
        continue
    fpath = os.path.join(txt_dir, fname)
    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into emails by ===== separator
    blocks = re.split(r'={5,}', content)
    
    for block in blocks:
        if not block.strip():
            continue
        # Find email name
        name_match = re.search(r'(?:^|\n)\s*(?:Email\s+name|Name)\s*[:\-]\s*(.+)', block, re.IGNORECASE)
        if not name_match:
            # Try to find any identifier
            first_line = block.strip().split('\n')[0].strip()
            ename = first_line[:60] if first_line else "???"
        else:
            ename = name_match.group(1).strip()

        # Find all promo codes in the block
        codes = re.findall(r'(?:Code|Код|Promo\s*code|promocode)[:\s]+([A-Z0-9_]+)', block, re.IGNORECASE)
        # Also check promocode_button fields
        btn_codes = re.findall(r'promocode_button[^:]*:\s*([A-Z0-9_]+)', block, re.IGNORECASE)
        codes.extend(btn_codes)
        
        # Deduplicate while preserving order
        seen = set()
        unique_codes = []
        for c in codes:
            cu = c.upper()
            if cu not in seen:
                seen.add(cu)
                unique_codes.append(c)

        if len(unique_codes) > 1:
            multi_bonus.append((fname, ename, unique_codes))

if multi_bonus:
    print(f"Found {len(multi_bonus)} emails with multiple promo codes:\n")
    for fname, ename, codes in multi_bonus:
        print(f"  [{fname}] {ename}")
        print(f"    Codes: {' / '.join(codes)}")
        print()
else:
    print("No multi-bonus emails found (besides already fixed tiered ones).")

wb_m.close()
wb_s.close()
