"""Full audit:
1. LIVE campaigns vs correct HH_Live master (from screenshot)
2. Excel anomalies: FS without game, weird words, empty cells
3. Locale consistency: all 4 locales must have same bonus & promo code
"""

import openpyxl, re, os, collections

SUMMARY = os.path.join("тексти", "хейхо", "HeyHo Emails Summary.xlsx")
TXT_DIR = os.path.join("тексти", "хейхо")

# ── Correct HH_Live master from screenshot ──────────────────────────
LIVE_MASTER = {
    "HOOK150":  "100%",
    "RAID145":  "90%",
    "MAP140":   "80%",
    "SABR135":  "75%",
    "KRKN130":  "70%",
    "RUM125":   "65%",
    "CREW120":  "60%",
    "FLAG115":  "50%",
    "KEEL110":  "45%",
    "BUCC105":  "40%",
    "SKLL100":  "35%",
    "REEF90":   "30%",
    "PLANK80":  "30%",
    "CHEST75":  "25%",
    "CUTLS70":  "25%",
    "CB4RUM":   "4% CB",
    "CB6MAP":   "6% CB",
    "CB9REEF":  "9% CB",
    "CB12SEA":  "12% CB",
    "CB15BAY":  "15% CB",
}

# ═══════════════════════════════════════════════════════════════════
# PART 1: LIVE audit — compare email bonus vs master bonus
# ═══════════════════════════════════════════════════════════════════
print("=" * 70)
print("PART 1: LIVE campaigns vs HH_Live master")
print("=" * 70)

wb = openpyxl.load_workbook(SUMMARY, data_only=True)
live_sheets = [s for s in wb.sheetnames if "LIVE" in s.upper()]

live_ok = 0
live_issues = []

for sname in live_sheets:
    ws = wb[sname]
    print(f"\n--- {sname} ---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        email = str(row[0]).strip()
        bonus = str(row[1]).strip() if row[1] else ""
        code  = str(row[2]).strip() if row[2] else ""
        game  = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if email == "Email" or code == "Promo Code":
            continue  # skip header dupes

        code_upper = code.upper()
        if code_upper not in LIVE_MASTER:
            live_issues.append((sname, email, f"Code '{code}' NOT in master"))
            print(f"  {email:25s} | code={code} NOT FOUND")
            continue

        master_bonus = LIVE_MASTER[code_upper]

        # Normalize for comparison
        email_bonus_norm = bonus.lower().replace(" ", "").replace("cashback", "cb")
        master_bonus_norm = master_bonus.lower().replace(" ", "")

        # For % bonuses: email says "100%" master says "100%"
        # For CB: email says "4% Cashback" or "4% CB", master says "4% CB"
        match = False
        if "cb" in master_bonus_norm:
            # Cashback comparison
            master_pct = re.search(r'(\d+)%', master_bonus_norm)
            email_pct = re.search(r'(\d+)%', email_bonus_norm)
            if master_pct and email_pct and master_pct.group(1) == email_pct.group(1):
                match = True
        else:
            # Deposit bonus comparison
            master_pct = re.search(r'(\d+)%', master_bonus_norm)
            email_pct = re.search(r'(\d+)%', email_bonus_norm)
            if master_pct and email_pct and master_pct.group(1) == email_pct.group(1):
                match = True

        if match:
            live_ok += 1
            print(f"  {email:25s} | {bonus:20s} | {code:12s} | OK")
        else:
            live_issues.append((sname, email, f"Bonus mismatch: email='{bonus}' vs master='{master_bonus}' (code={code})"))
            print(f"  {email:25s} | {bonus:20s} | {code:12s} | *** MISMATCH *** master={master_bonus}")

print(f"\n{'='*50}")
print(f"LIVE Summary: OK={live_ok}, Issues={len(live_issues)}")
for s, e, msg in live_issues:
    print(f"  [{s}] {e}: {msg}")

# ═══════════════════════════════════════════════════════════════════
# PART 2: Excel anomalies across ALL sheets
# ═══════════════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("PART 2: Excel anomalies (FS without game, weird values, empty cells)")
print("=" * 70)

anomalies = []

for sname in wb.sheetnames:
    if sname == "Summary":
        continue
    ws = wb[sname]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        email = str(row[0]).strip()
        bonus = str(row[1]).strip() if row[1] else ""
        code  = str(row[2]).strip() if row[2] else ""
        game  = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if email == "Email" or code == "Promo Code":
            continue

        # Check: FS bonus but no game
        if re.search(r'\d+\s*FS|\d+\s*Free\s*Spin', bonus, re.IGNORECASE) and not game:
            anomalies.append((sname, email, f"Has FS bonus '{bonus}' but no game"))

        # Check: empty bonus
        if not bonus:
            anomalies.append((sname, email, "Empty bonus"))

        # Check: empty promo code
        if not code:
            anomalies.append((sname, email, f"Empty promo code (bonus='{bonus}')"))

        # Check: game but no FS in bonus
        if game and not re.search(r'FS|Free\s*Spin|free spin', bonus, re.IGNORECASE):
            # Some games may be for deposit bonuses too, just flag
            if "%" not in bonus and "CB" not in bonus.upper() and "cashback" not in bonus.lower():
                anomalies.append((sname, email, f"Has game '{game}' but bonus '{bonus}' doesn't mention FS or %"))

        # Check: bonus contains suspicious words (not a real bonus)
        if bonus and not re.search(r'\d', bonus) and bonus not in ("Bonus",):
            anomalies.append((sname, email, f"Bonus has no numbers: '{bonus}'"))

        # Check: promo code looks weird (too short, has spaces)
        if code and (len(code) < 3 or " " in code.strip()):
            if "/" not in code:  # skip tiered codes with /
                anomalies.append((sname, email, f"Suspicious promo code: '{code}'"))

        # Check: game contains weird text
        if game and len(game) > 50:
            anomalies.append((sname, email, f"Game name too long: '{game[:60]}...'"))

        # Check: "NEEDED" or placeholder
        if "NEEDED" in bonus.upper() or "NEEDED" in code.upper() or "NEEDED" in game.upper():
            anomalies.append((sname, email, f"Contains 'NEEDED' placeholder"))
        if "TBD" in bonus.upper() or "TBD" in code.upper():
            anomalies.append((sname, email, f"Contains 'TBD' placeholder"))

if anomalies:
    print(f"\nFound {len(anomalies)} anomalies:\n")
    for s, e, msg in anomalies:
        print(f"  [{s}] {e}: {msg}")
else:
    print("\nNo anomalies found.")

# ═══════════════════════════════════════════════════════════════════
# PART 3: Locale consistency — check .txt source files
# ═══════════════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("PART 3: Locale consistency (all locales must have same bonus & code)")
print("=" * 70)

locale_issues = []

for fname in sorted(os.listdir(TXT_DIR)):
    if not fname.endswith(".txt"):
        continue

    fpath = os.path.join(TXT_DIR, fname)
    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into email blocks
    blocks = re.split(r'={5,}', content)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        # Parse rows: key: value
        rows = {}
        for line in block.split('\n'):
            line = line.strip()
            if ':' in line:
                key, _, val = line.partition(':')
                key = key.strip().lower()
                val = val.strip()
                if key not in rows:
                    rows[key] = val
                else:
                    # Multiple values for same key -- append
                    if isinstance(rows[key], list):
                        rows[key].append(val)
                    else:
                        rows[key] = [rows[key], val]

        # Skip if no name
        name_key = None
        for k in rows:
            if 'name' in k and 'email' in k:
                name_key = k
                break
            elif k == 'name':
                name_key = k
                break
        if not name_key:
            continue

        email_name = rows[name_key] if isinstance(rows[name_key], str) else rows[name_key][0]

        # Find locale field
        locale_val = rows.get('locale', '')
        if isinstance(locale_val, list):
            locale_val = locale_val[0]

        # Find all promo codes in the block
        promo_codes = []
        promo_btn_pattern = re.compile(r'promocode_button', re.IGNORECASE)
        for line in block.split('\n'):
            line_stripped = line.strip()
            if ':' in line_stripped:
                key_part = line_stripped.split(':')[0].strip().lower()
                val_part = line_stripped.split(':', 1)[1].strip()
                if 'promocode' in key_part and 'button' in key_part:
                    if val_part and val_part.upper() not in ('', 'NONE', 'NULL'):
                        promo_codes.append(val_part)

        # Find bonus-like text fields
        bonus_texts = []
        for line in block.split('\n'):
            line_stripped = line.strip()
            if ':' in line_stripped:
                key_part = line_stripped.split(':')[0].strip().lower()
                val_part = line_stripped.split(':', 1)[1].strip()
                # Look for text fields that mention bonus/FS/cashback
                if re.match(r'text_\d+', key_part):
                    if re.search(r'\d+%|\d+\s*FS|\d+\s*free\s*spin|cashback', val_part, re.IGNORECASE):
                        bonus_texts.append(val_part)

    # Now do a proper per-email locale comparison
    # Re-parse: group by email name, collect locale data
    for fname2 in sorted(os.listdir(TXT_DIR)):
        break  # we'll do this differently below

# Better approach: parse each file, group rows by email name
# Each .txt file has blocks separated by =====
# Within each block, multiple rows with "Email name:", "locale:", "promocode_button_1:" etc.

print("\nRe-parsing all .txt files for locale comparison...\n")

locale_issues = []

for fname in sorted(os.listdir(TXT_DIR)):
    if not fname.endswith(".txt"):
        continue

    fpath = os.path.join(TXT_DIR, fname)
    with open(fpath, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into email blocks by ===== separator
    blocks = re.split(r'={5,}', content)

    # Group blocks by email name
    email_groups = collections.defaultdict(list)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        lines = block.split('\n')
        data = {}
        for line in lines:
            line = line.strip()
            if not line or ':' not in line:
                continue
            key, _, val = line.partition(':')
            key = key.strip()
            val = val.strip()
            key_lower = key.lower()
            if key_lower not in data:
                data[key_lower] = val
            # Collect all promocode_button fields
            if 'promocode_button' in key_lower:
                if 'promo_codes' not in data:
                    data['promo_codes'] = []
                if val:
                    data['promo_codes'].append(val.upper())

        # Find email name
        ename = None
        for k in ('email name', 'name'):
            if k in data:
                ename = data[k]
                break
        if not ename:
            continue

        locale = data.get('locale', 'Default')
        promo_codes = data.get('promo_codes', [])

        email_groups[ename].append({
            'locale': locale,
            'promo_codes': promo_codes,
            'data': data,
        })

    # Now compare locales within each email
    for ename, locales in email_groups.items():
        if len(locales) <= 1:
            continue  # only 1 locale, nothing to compare

        # Get reference (first locale)
        ref = locales[0]
        ref_codes = sorted(ref['promo_codes'])

        for loc in locales[1:]:
            loc_codes = sorted(loc['promo_codes'])
            if loc_codes != ref_codes:
                locale_issues.append((
                    fname, ename,
                    f"Promo code mismatch: {ref['locale']}={ref_codes} vs {loc['locale']}={loc_codes}"
                ))

if locale_issues:
    print(f"Found {len(locale_issues)} locale inconsistencies:\n")
    for f, e, msg in locale_issues:
        print(f"  [{f}] {e}: {msg}")
else:
    print("All locales consistent for promo codes.")

wb.close()
print(f"\n{'='*70}")
print("AUDIT COMPLETE")
print("=" * 70)
