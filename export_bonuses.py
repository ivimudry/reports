import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ---- Read original ----
wb = openpyxl.load_workbook('C:/Projects/REPORTS/Book1.xlsx')
ws = wb.active

rows = []
for r in range(2, ws.max_row + 1):
    c1 = ws.cell(row=r, column=1).value
    if c1 is None:
        continue
    section = str(c1).strip()
    bonus_type = str(ws.cell(row=r, column=2).value or '').strip()
    active = str(ws.cell(row=r, column=3).value or '').strip()
    details = str(ws.cell(row=r, column=4).value or '')
    conditions = str(ws.cell(row=r, column=5).value or '')

    # Parse title
    title_m = re.search(r'Title:\s*(.+?)(?:\s*\(Input with identifier)', details)
    title = title_m.group(1).strip() if title_m else ''

    # Parse identifier
    ident_m = re.search(r'Input with identifier:\s*(\S+)\)', details)
    identifier = ident_m.group(1).strip() if ident_m else ''

    # Parse FS count
    fs_m = re.search(r'Freespins count:\s*(\d+)', details)
    fs_count = int(fs_m.group(1)) if fs_m else None

    # Parse amount %
    amt_m = re.search(r'Amount\s+([\d.]+)%', details)
    amount_pct = amt_m.group(1) + '%' if amt_m else None

    # Parse fixed amount (for cash bonuses)
    fixed_m = re.search(r'Amount per currency\s+([\d,.]+)\s*ZAR', details)
    fixed_amount = fixed_m.group(1) + ' ZAR' if fixed_m else None

    # Parse max amount
    max_m = re.search(r'Maximum amount per currency\s+([\d,.]+)\s*ZAR', details)
    max_amount = max_m.group(1) + ' ZAR' if max_m else None

    # Parse games
    games_m = re.search(r'Games identifiers:\s*\[([^\]]+)\]', details)
    if games_m:
        raw_games = games_m.group(1)
        game_names = re.findall(r'"[^"]*:([^"]+)"', raw_games)
        if not game_names:
            game_names = re.findall(r'"([^"]+)"', raw_games)
        game = ', '.join(game_names)
    else:
        game = ''

    # Parse wager
    wager_m = re.search(r'Wager Multiplier:\s*(\d+)', details)
    wager = int(wager_m.group(1)) if wager_m else None

    # Parse bonus code
    code_m = re.search(r'Bonus code:\s*\[([^\]]+)\]', conditions)
    if code_m:
        codes_raw = code_m.group(1)
        codes_list = re.findall(r'"([^"]+)"', codes_raw)
        # Get the uppercase one or first
        promo_code = codes_list[0] if codes_list else ''
        for c in codes_list:
            if c.isupper():
                promo_code = c
                break
    else:
        promo_code = ''

    # Parse min deposit
    min_dep_m = re.search(r'Min value:\s*([\d,.]+)\s*ZAR', conditions)
    min_deposit = min_dep_m.group(1) + ' ZAR' if min_dep_m else None
    
    # Parse deposit range from conditions
    dep_range_m = re.search(r'Value should be greater or equal to\s*([\d,.]+)\s*ZAR\s*and less than\s*([\d,.]+)\s*ZAR', conditions)
    deposit_range = f"{dep_range_m.group(1)}-{dep_range_m.group(2)} ZAR" if dep_range_m else None

    # Parse frequency
    freq_m = re.search(r'Frequency:\s*no more than\s*(\d+)\s*(?:in\s*(\d+)\s*(\w+))?\s*among\s*(\S+)', conditions)
    frequency = ''
    if freq_m:
        count = freq_m.group(1)
        period_num = freq_m.group(2)
        period_unit = freq_m.group(3)
        if period_num and period_unit:
            frequency = f"{count} per {period_num} {period_unit}"
        else:
            frequency = f"{count} total"

    # Parse available dates
    avail_m = re.search(r'Available from\s*(\S+)\s*until\s*(\S+)', conditions)
    available = f"{avail_m.group(1)} - {avail_m.group(2)}" if avail_m else ''

    # Determine bonus value display
    if amount_pct:
        bonus_value = amount_pct
    elif fixed_amount:
        bonus_value = fixed_amount
    elif fs_count:
        bonus_value = f"{fs_count} FS"
    else:
        bonus_value = ''

    rows.append({
        'row_num': r,
        'section': section,
        'bonus_type': bonus_type,
        'active': active,
        'title': title,
        'identifier': identifier,
        'bonus_value': bonus_value,
        'fs_count': fs_count,
        'amount_pct': amount_pct,
        'fixed_amount': fixed_amount,
        'max_amount': max_amount,
        'game': game,
        'promo_code': promo_code,
        'wager': wager,
        'min_deposit': min_deposit or '',
        'deposit_range': deposit_range or '',
        'frequency': frequency,
        'available': available,
    })

print(f"Parsed {len(rows)} bonus rows")

# ---- Find duplicates ----
# Group by title (cleaned of emoji)
def clean_title(t):
    return re.sub(r'[^\w\s]', '', t).strip().lower()

title_groups = defaultdict(list)
for i, row in enumerate(rows):
    key = clean_title(row['title'])
    if key:
        title_groups[key].append(i)

# Mark differences for duplicates
for i, row in enumerate(rows):
    row['difference'] = ''

for key, indices in title_groups.items():
    if len(indices) <= 1:
        continue
    # Find what differs between them
    base = rows[indices[0]]
    for idx in indices[1:]:
        other = rows[idx]
        diffs = []
        for field in ['section', 'bonus_value', 'game', 'promo_code', 'min_deposit', 'deposit_range', 'wager', 'frequency', 'available', 'identifier']:
            v1 = str(base.get(field, ''))
            v2 = str(other.get(field, ''))
            if v1 != v2:
                diffs.append(f"{field}: {v1} vs {v2}")
        diff_text = '; '.join(diffs) if diffs else 'Exact duplicate'
        # Mark both
        if not rows[indices[0]]['difference']:
            rows[indices[0]]['difference'] = f"Duplicate #{len(indices)} - see row differences"
        rows[idx]['difference'] = '; '.join(diffs) if diffs else 'Exact duplicate'

# Also check by identifier base (without numeric suffix) for related bonuses
# These aren't true duplicates but related tiers

# ---- Write to new xlsx ----
out = openpyxl.Workbook()
ows = out.active
ows.title = "PantherBet Bonuses"

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Calibri', size=10)
data_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# Section colors
section_fills = {
    'Registration': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'Coupon Input': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'Deposit': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    'Loyalty': PatternFill(start_color='E4D5F0', end_color='E4D5F0', fill_type='solid'),
    'Scheduler': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'Random': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
}

dup_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Headers
headers = ['#', 'Section', 'Type', 'Active', 'Title', 'Bonus Value', 'Game', 'Promo Code', 
           'Wager', 'Min Deposit / Range', 'Frequency', 'Available', 'Identifier', 'Difference']

for c, h in enumerate(headers, 1):
    cell = ows.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data
for i, row in enumerate(rows):
    r = i + 2  # Excel row
    
    dep_info = row['min_deposit'] or row['deposit_range'] or ''
    
    values = [
        i + 1,
        row['section'],
        row['bonus_type'],
        row['active'],
        row['title'],
        row['bonus_value'],
        row['game'],
        row['promo_code'],
        row['wager'] if row['wager'] else '',
        dep_info,
        row['frequency'],
        row['available'],
        row['identifier'],
        row['difference'],
    ]
    
    for c, v in enumerate(values, 1):
        cell = ows.cell(row=r, column=c, value=v)
        cell.font = data_font
        cell.alignment = center_align if c in (1, 3, 4, 8, 9) else data_align
        cell.border = thin_border
    
    # Section color
    sec = row['section']
    for section_key, fill in section_fills.items():
        if section_key.lower() in sec.lower():
            for c in range(1, len(headers) + 1):
                ows.cell(row=r, column=c).fill = fill
            break
    
    # Highlight duplicates
    if row['difference']:
        for c in range(1, len(headers) + 1):
            ows.cell(row=r, column=c).fill = dup_fill

# Column widths
widths = [4, 14, 10, 7, 40, 14, 30, 14, 7, 20, 16, 24, 35, 50]
for c, w in enumerate(widths, 1):
    ows.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# Freeze header row
ows.freeze_panes = 'A2'

# Auto filter
ows.auto_filter.ref = f"A1:N{len(rows) + 1}"

output_path = 'C:/Projects/REPORTS/PantherBet_Bonuses.xlsx'
out.save(output_path)
print(f"Saved to {output_path}")
print(f"Total rows: {len(rows)} (excluding header)")

# Print duplicate summary
dup_count = sum(1 for r in rows if r['difference'])
print(f"Rows with duplicate notes: {dup_count}")
for r in rows:
    if r['difference']:
        print(f"  Row {r['row_num']}: {r['title'][:50]} -> {r['difference'][:80]}")
