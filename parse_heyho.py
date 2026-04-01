"""Parse all HeyHo Casino Table data files and generate Excel summary."""

import os
import re
import glob

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def strip_html(html_text):
    """Remove HTML tags, decode entities, normalize whitespace."""
    if not html_text:
        return ""
    text = re.sub(r'<br\s*/?>', ' ', html_text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def parse_file(filepath):
    """Parse a Table data file into list of email records (Default locale only)."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    records = []
    current = {}

    for line in lines:
        m = re.match(r'^(\w[\w_]*?):\s*(.*)', line)
        if m:
            key, val = m.group(1), m.group(2).strip()
            if key == 'name' and current:
                records.append(current)
                current = {}
            current[key] = val
        elif line.strip() == '' and current:
            pass  # blank line inside record
    if current:
        records.append(current)

    # Filter only Default locale, deduplicate by name
    seen = set()
    defaults = []
    for rec in records:
        if rec.get('locale', '') == 'Default' and rec.get('name', '') not in seen:
            seen.add(rec['name'])
            defaults.append(rec)

    return defaults


KNOWN_GAMES = [
    'Gates of Olympus 1000', 'Gates of Olympus',
    'Sweet Bonanza 1000', 'Sweet Bonanza Xmas', 'Sweet Bonanza',
    'Sugar Rush 1000', 'Sugar Rush',
    'Big Bass Splash', 'Big Bass Bonanza', 'Big Bass',
    'Book of Dead',
    'Hand of Midas 2', 'Hand of Midas',
    'Starlight Princess 1000', 'Starlight Princess',
    'Wolf Gold',
    'Fruit Party 2', 'Fruit Party',
    'Razor Shark',
    'Fire Joker',
    'Starburst XXXtreme', 'Starburst',
    'Legacy of Dead',
    'Rise of Olympus',
    'Reactoonz 2', 'Reactoonz',
    'The Dog House Megaways', 'The Dog House',
    'Buffalo King Megaways', 'Buffalo King',
    'Eye of Horus',
    'Wanted Dead or a Wild',
    'Wild West Gold',
    'Mustang Gold',
    "Jammin' Jars 2", "Jammin' Jars",
    'Gold Rush',
    'Madame Destiny Megaways', 'Madame Destiny',
    'Gems Bonanza',
    'Power of Thor Megaways',
    'Great Rhino Megaways', 'Great Rhino',
    'Floating Dragon',
    'Might of Ra',
    'Book of Fallen',
    'Lucky Fisherman',
    'Shining Crown',
    'Cleocatra',
    'Candy Boom',
    'Oink Oink Oink',
    'Le Bandit',
    'Zeus vs Hades',
    'Wisdom of Athena',
    'Golden Glyph 3', 'Golden Glyph',
    'Thunder Screech',
    'Sizzling Eggs',
    'Queenie',
    '5 Lions Megaways',
    'Piranha Pays',
]


def extract_game(search_text):
    """Extract game name from text using known game list + fallback patterns."""
    # First try exact known game names (longest first to match "Sugar Rush 1000" before "Sugar Rush")
    for g in KNOWN_GAMES:
        if g.lower() in search_text.lower():
            return g

    # Noise words that are definitely NOT game names
    noise = {'ahoy', 'your', 'next', 'deposit', 'this', 'wild', 'the', 'a', 'an',
             'into', 'account', 'second', 'third', 'first', 'brave', 'ready', 'launch',
             'await', 'last', 'week', 'extra', 'credit', 'extends', 'play', 'clean',
             'boost', 'thank', 'you', 'with'}

    def _is_valid_game(val):
        words = val.lower().split()
        if len(words) == 0:
            return False
        if all(w in noise for w in words):
            return False
        if len(val) > 40:
            return False
        return True

    # Fallback: "on/in/for [Capitalized Game Name]" after spins context
    m = re.search(
        r'(?:Spins?|FS)\s+(?:on|in|for|at)\s+([A-Z][A-Za-z0-9\s\'\-:]+?)(?:\s*[.!,;?\n]|\s+into|\s+on\s)',
        search_text
    )
    if m:
        val = m.group(1).strip().rstrip('.')
        if _is_valid_game(val):
            return val

    # Subject: "XX FS on GameName"
    m = re.search(
        r'(?:FS|Free Spins?)\s+(?:on|in|for)\s+(.+?)$',
        search_text.split('\n')[0], re.IGNORECASE
    )
    if m:
        val = m.group(1).strip().rstrip('.').rstrip('!')
        if _is_valid_game(val):
            return val

    return ''


def extract_info(rec):
    """Extract bonus, promo code, game name from a record."""
    name = rec.get('name', '')
    subject = strip_html(rec.get('subject', ''))
    preheader = strip_html(rec.get('preheader', ''))

    # Collect all text fields
    all_texts = []
    for k, v in rec.items():
        if k.startswith('text_'):
            all_texts.append(strip_html(v))
    full_text = ' '.join(all_texts)

    # Promo code: prefer promocode_button_1, then search in texts
    promo = rec.get('promocode_button_1', '').strip()
    if not promo:
        promo_matches = re.findall(
            r'\b(?:code\s+|enter\s+|gib\s+|voer\s+.*?in\s+)([A-Z][A-Z0-9]{3,})\b',
            full_text, re.IGNORECASE
        )
        if promo_matches:
            promo = promo_matches[0].upper()

    # Game name
    search_text = subject + ' ' + full_text
    game = extract_game(search_text)

    # Bonus extraction
    bonus = ''
    bonus_parts = []

    # Cash Back / Cashback
    cb_match = re.search(r'(\d+)%\s*Cash\s*back', search_text, re.IGNORECASE)
    if cb_match:
        bonus_parts = [f"{cb_match.group(1)}% Cashback"]
    else:
        # Percentage bonus (deposit match) — look for patterns like "120%", "get 120%"
        pct_match = re.search(
            r'(?:get|grab|claim|deposit.*?|unlock|receive|revealed)?\s*(\d{2,3})%(?!\s*Cash)',
            search_text, re.IGNORECASE
        )
        if pct_match:
            pct_val = pct_match.group(1)
            if int(pct_val) >= 50:  # likely a deposit bonus, not a small number
                bonus_parts.append(f"{pct_val}%")

        # Free spins
        fs_match = re.search(r'(\d+)\s*(?:Free Spins?|FS|Freispiele|tours gratuits|gratis spins)', search_text, re.IGNORECASE)
        if fs_match:
            bonus_parts.append(f"{fs_match.group(1)} FS")

    # "up to" bonus amounts
    upto_match = re.search(r'(?:up to|bis zu|tot)\s*[€]?\s*(\d[\d,]*)', search_text, re.IGNORECASE)

    if bonus_parts:
        bonus = ' + '.join(bonus_parts)
        if upto_match and 'Cashback' not in bonus:
            bonus += f" (up to €{upto_match.group(1)})"
    else:
        # No standard bonus — short description from subject
        desc = subject
        desc = re.sub(r'[\U0001F300-\U0001FFFF\u200D\u2600-\u26FF\u2700-\u27BF\uFE0F]', '', desc).strip()
        if desc:
            bonus = desc
        else:
            bonus = preheader or '—'

    return {
        'name': name,
        'bonus': bonus,
        'promo': promo,
        'game': game
    }


def main():
    folder = r"C:\Projects\REPORTS\тексти\хейхо"
    files = sorted(glob.glob(os.path.join(folder, "*.txt")))

    if not files:
        print("No files found!")
        return

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    campaign_font = Font(name='Calibri', bold=True, size=12, color='2F5496')

    total_emails = 0

    for filepath in files:
        fname = os.path.basename(filepath).replace(' - Table data.txt', '')
        print(f"\nProcessing: {fname}")

        records = parse_file(filepath)
        print(f"  Found {len(records)} unique emails")
        total_emails += len(records)

        # Sheet name max 31 chars
        sheet_name = fname[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Campaign title row
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = fname
        title_cell.font = campaign_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Headers
        headers = ['Email', 'Bonus', 'Promo Code', 'Game']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, rec in enumerate(records, 3):
            info = extract_info(rec)
            values = [info['name'], info['bonus'], info['promo'], info['game']]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = cell_align
                cell.border = thin_border
                if (row_idx - 3) % 2 == 1:
                    cell.fill = alt_fill

            print(f"  {info['name']:15s} | {info['bonus']:35s} | {info['promo']:12s} | {info['game']}")

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 30

        # Freeze header
        ws.freeze_panes = 'A3'

    # Summary sheet at the beginning
    summary = wb.create_sheet(title="Summary", index=0)
    summary.merge_cells('A1:C1')
    summary['A1'].value = 'HeyHo Casino — Email Summary'
    summary['A1'].font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    summary.row_dimensions[1].height = 35

    sum_headers = ['Campaign', 'Emails', 'Sheet']
    for col_idx, h in enumerate(sum_headers, 1):
        cell = summary.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, filepath in enumerate(files, 3):
        fname = os.path.basename(filepath).replace(' - Table data.txt', '')
        records = parse_file(filepath)
        sheet_name = fname[:31]

        summary.cell(row=row_idx, column=1, value=fname).border = thin_border
        summary.cell(row=row_idx, column=2, value=len(records)).border = thin_border
        summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
        link_cell = summary.cell(row=row_idx, column=3, value=sheet_name)
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color='0563C1', underline='single')
        link_cell.border = thin_border

        if (row_idx - 3) % 2 == 1:
            for c in range(1, 4):
                summary.cell(row=row_idx, column=c).fill = alt_fill

    # Total row
    total_row = len(files) + 3
    summary.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    summary.cell(row=total_row, column=1).border = thin_border
    summary.cell(row=total_row, column=2, value=total_emails).font = Font(bold=True)
    summary.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
    summary.cell(row=total_row, column=2).border = thin_border

    summary.column_dimensions['A'].width = 35
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 35
    summary.freeze_panes = 'A3'

    output_path = os.path.join(folder, "HeyHo Emails Summary.xlsx")
    wb.save(output_path)
    print(f"\n✅ Excel saved: {output_path}")
    print(f"   Total: {total_emails} unique emails across {len(files)} campaigns")


if __name__ == '__main__':
    main()
