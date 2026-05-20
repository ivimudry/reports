from openpyxl import load_workbook
import sys
sys.stdout.reconfigure(encoding='utf-8')
wb = load_workbook(r"c:\Projects\CLIENTS\CHEEL4REEL\інше\Loyalty_Program_Adjustment_C4R.xlsx", data_only=True)
ws = wb["Loyalty Program"]
maxcol = 0
for row in ws.iter_rows(values_only=True):
    for j,v in enumerate(row):
        if v is not None and str(v).strip(): maxcol = max(maxcol, j+1)
print(f"maxcol={maxcol}")
for i,row in enumerate(ws.iter_rows(values_only=True),1):
    trimmed = row[:maxcol]
    if not any(v is not None and str(v).strip() for v in trimmed): continue
    print(f"R{i}: " + " | ".join("" if v is None else str(v).replace("\n"," / ") for v in trimmed))
