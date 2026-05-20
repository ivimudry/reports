from openpyxl import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
for path in [r'c:\Projects\CLIENTS\CHEEL4REEL\інше\Loyalty_Program_Adjustment_C4R.xlsx', r'c:\Projects\CLIENTS\CHEEL4REEL\інше\FS - Games.xlsx']:
    print('='*80); print('FILE:', path); print('='*80)
    wb = load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        # find true bounds by scanning
        rows = list(ws.iter_rows(values_only=True))
        # trim empty trailing rows
        while rows and all(v is None or (isinstance(v,str) and v.strip()=='') for v in rows[-1]):
            rows.pop()
        # find max col with any data
        maxc = 0
        for r in rows:
            for i,v in enumerate(r):
                if v is not None and not (isinstance(v,str) and v.strip()==''):
                    if i+1 > maxc: maxc = i+1
        print(f'\n--- SHEET: {sn} (effective rows={len(rows)} cols={maxc}) ---')
        for ri, r in enumerate(rows, 1):
            trimmed = r[:maxc]
            print(f'R{ri}: ' + ' | '.join('' if v is None else str(v) for v in trimmed))
